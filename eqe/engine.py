from .utils import excel_ref_to_index, parse_excel_range, add_offset, column_letter_to_index
from pathlib import Path


class ExcelQueryEngine:
    def __init__(self, excel_source):
        """Create an engine from either an in-memory dict or an Excel filename.

        - If `excel_source` is a `str` or `pathlib.Path`, it is treated as a
          filename and the workbook is loaded (requires `openpyxl`). Each sheet
          is converted to a list-of-lists (rows as lists, values_only).
        - If `excel_source` is a dict (mapping sheet name -> list-of-rows), it
          is used directly for backward compatibility.
        """
        # If a path-like object or string: load workbook and build data dict
        if isinstance(excel_source, (str, Path)):
            path = Path(excel_source)
            try:
                from openpyxl import load_workbook
            except ModuleNotFoundError as exc:
                raise ModuleNotFoundError(
                    "openpyxl is required to load Excel files. "
                    "Install with 'pip install openpyxl'.") from exc

            wb = load_workbook(filename=str(path), data_only=True)
            data = {}
            for sheet in wb.worksheets:
                # iter_rows with values_only returns tuples per row
                rows = [list(r) for r in sheet.iter_rows(values_only=True)]
                data[sheet.title] = rows

            self.data = data
        else:
            # Assume user supplied already-parsed data structure
            self.data = excel_source

    def get_cell(self, sheet, row, col):
        return self.data.get(sheet, [])[row][col]

    def find_by_value(self, sheet, target):
        matches = []
        sheet_data = self.data.get(sheet, [])
        for row_idx, row in enumerate(sheet_data):
            for col_idx, value in enumerate(row):
                if value == target:
                    matches.append((row_idx, col_idx))
        return matches

    def get_adjacent_value(self, sheet, search_value, offset=(0, 1)):
        matches = self.find_by_value(sheet, search_value)
        if not matches:
            return None
        row, col = matches[0]
        adj_row = row + offset[0]
        adj_col = col + offset[1]
        return self.data[sheet][adj_row][adj_col]

    def extract_table_from_header(self, sheet, header_row_idx):
        sheet_data = self.data.get(sheet, [])
        headers = sheet_data[header_row_idx]
        table_rows = sheet_data[header_row_idx + 1 :]
        return [dict(zip(headers, row)) for row in table_rows if any(row)]

    def get_range(self, sheet, start, end=None, size=None):
        result = []
        start_row, start_col = start

        if end:
            end_row, end_col = end
        elif size:
            end_row = start_row + size[0] - 1
            end_col = start_col + size[1] - 1
        else:
            raise ValueError("You must provide either 'end' or 'size'.")

        for row_idx in range(start_row, end_row + 1):
            row_values = []
            for col_idx in range(start_col, end_col + 1):
                value = self.get_cell(sheet, row_idx, col_idx)
                row_values.append(value)
            result.append(row_values)

        return result

    @staticmethod
    def add_offset(base, offset):
        return add_offset(base, offset)

    @staticmethod
    def excel_ref_to_index(cell_ref):
        return excel_ref_to_index(cell_ref)

    @staticmethod
    def parse_excel_range(range_ref):
        return parse_excel_range(range_ref)

    def get_range_by_ref(self, sheet, ref):
        start, end = self.parse_excel_range(ref)
        return self.get_range(sheet, start=start, end=end)

    def get_columns_from_row(self, sheet, columns, start_row):
        result = []
        sheet_data = self.data.get(sheet, [])
        start_idx = start_row - 1
        col_indices = [column_letter_to_index(col) for col in columns]

        for row in sheet_data[start_idx:]:
            selected = []
            for col_idx in col_indices:
                if col_idx < len(row):
                    selected.append(row[col_idx])
                else:
                    selected.append(None)
            result.append(selected)

        return result

    def get_columns_from_row_to_row(self, sheet, columns, start_row, end_row):
        result = []
        sheet_data = self.data.get(sheet, [])
        start_idx = start_row - 1
        end_idx = end_row - 1
        col_indices = [column_letter_to_index(col) for col in columns]

        for row in sheet_data[start_idx:end_idx]:
            selected = []
            for col_idx in col_indices:
                if col_idx < len(row):
                    selected.append(row[col_idx])
                else:
                    selected.append(None)
            result.append(selected)

        return result
